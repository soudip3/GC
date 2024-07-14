from openpyxl import load_workbook
from getRolePermission import getRolePermission

wb = load_workbook('roles.xlsx')
ws = wb.active

max_row = ws.max_row

for i in range(max_row-1):
    getRolePermission(ws.cell(i+2,1).value)