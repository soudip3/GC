import sys
from openpyxl import Workbook,load_workbook

wb = Workbook()

def loadExcel(response_json,role):
    ws = wb.create_sheet(f"Role_{role}")
    permissionPoliciesLen = len(response_json["entities"][0]["permissionPolicies"])
    k=1
    for i in range(permissionPoliciesLen):
        actionSetLen = len(response_json["entities"][0]["permissionPolicies"][i]["actionSet"])
        for j in range (actionSetLen):
            domain = response_json["entities"][0]["permissionPolicies"][i]["domain"]
            entityName = response_json["entities"][0]["permissionPolicies"][i]["entityName"]
            actionSet = response_json["entities"][0]["permissionPolicies"][i]["actionSet"][j]
            if actionSet == "*":
                actionSet = "ALL"
            if entityName == "*":
                entityName = "ALL"
            permission = domain+" > "+entityName+" > "+actionSet
            ws[f"A{k}"] = permission
            k=k+1
    wb.save("rolesExtraction.xlsx")
